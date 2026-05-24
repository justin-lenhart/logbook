"""
Legacy logbook import — one Airtable Flights row per aircraft type, aggregated totals.

Reads the Master sheet from an AnytimeLogbooks .xlsx export and upserts summary
rows into the Logbook Airtable base. Aircraft records are created on demand if
they don't already exist.

Default: dry-run (nothing is written). Pass --commit to execute.

Usage:
    uv run python scripts/import_legacy.py [--xlsx PATH] [--commit]
"""

from __future__ import annotations

import argparse
import datetime
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from pyairtable import Api

# ── Paths ─────────────────────────────────────────────────────────────────────

TOOLS_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = TOOLS_ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "misc" / "LEN2J-AnytimeLogbook2025.12.01.xlsx"
ENV_PATH = TOOLS_ROOT / ".env"

# ── Airtable IDs ──────────────────────────────────────────────────────────────

BASE_ID      = "appJSWcRpFXhwWOqd"
TBL_FLIGHTS  = "tblKi49d48CaDONVO"
TBL_AIRCRAFT = "tblsIsMVU2cxiewr7"

# ── Master sheet column indices (0-based) ─────────────────────────────────────

COL_DATE       = 1   # B — flight date
COL_TYPE       = 2   # C — aircraft type code
COL_TOTAL      = 5   # F — total block time (hours)
COL_DAY_LDG    = 13  # N — day landings
COL_NIGHT_LDG  = 14  # O — night landings
COL_NIGHT_TIME = 15  # P — night flight time
COL_ACT_IN     = 16  # Q — actual instrument time
COL_XC         = 19  # T — cross-country time
COL_PIC        = 23  # X — PIC time
COL_SIC        = 24  # Y — SIC time
COL_DUAL       = 25  # Z — dual received
COL_CFI        = 26  # AA — dual given (IP/CFI)

# ── Aircraft metadata for records not yet in Airtable ─────────────────────────
# (code, faa_type, category, class_or_none, engine_type)
# All fields are singleSelect in Airtable; typecast=True creates new options.
AIRCRAFT_META: dict[str, tuple[str, str, str | None, str]] = {
    "AH1W":  ("AH-1",       "Rotorcraft",   "Helicopter", "Turboshaft"),
    "AH1Z":  ("AH-1",       "Rotorcraft",   "Helicopter", "Turboshaft"),
    "TH57B": ("TH-57",      "Rotorcraft",   "Helicopter", "Turboshaft"),
    "TH57C": ("TH-57",      "Rotorcraft",   "Helicopter", "Turboshaft"),
    "T6B":   ("T-6B",       "Airplane",     "Single-engine Land", "Turboprop"),
    "C172":  ("C172",       "Airplane",     "Single-engine Land", "Reciprocating"),
    "PA44":  ("PA-44",      "Airplane",     "Multi-engine Land",  "Reciprocating"),
    "UC12W": ("UC-12W",     "Airplane",     "Multi-engine Land",  "Turboprop"),
    "UH1Y":  ("UH-1Y",     "Rotorcraft",   "Helicopter", "Turboshaft"),
    "MV22B": ("MV-22B",    "Powered Lift", None,         "Turboshaft"),
}

SOURCE_FILENAME = "LEN2J-AnytimeLogbook2025.12.01.xlsx"


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _num(v: object) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def aggregate_master(ws) -> dict[str, dict]:
    """Aggregate per-aircraft-type totals from the Master sheet (data from row 6)."""
    ac: dict[str, dict] = defaultdict(lambda: {
        "total": 0.0, "day_ldg": 0.0, "night_ldg": 0.0, "night_time": 0.0,
        "act_in": 0.0, "xc": 0.0, "pic": 0.0, "sic": 0.0,
        "dual": 0.0, "cfi": 0.0, "dates": [], "rows": 0,
    })
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        date_val = row[COL_DATE]
        type_val = row[COL_TYPE]
        if not isinstance(date_val, datetime.datetime) or not isinstance(type_val, str):
            continue
        d = ac[type_val]
        d["rows"]       += 1
        d["dates"].append(date_val)
        d["total"]      += _num(row[COL_TOTAL])
        d["day_ldg"]    += _num(row[COL_DAY_LDG])
        d["night_ldg"]  += _num(row[COL_NIGHT_LDG])
        d["night_time"] += _num(row[COL_NIGHT_TIME])
        d["act_in"]     += _num(row[COL_ACT_IN])
        d["xc"]         += _num(row[COL_XC])
        d["pic"]        += _num(row[COL_PIC])
        d["sic"]        += _num(row[COL_SIC])
        d["dual"]       += _num(row[COL_DUAL])
        d["cfi"]        += _num(row[COL_CFI])
    return dict(ac)


# ── Record builders ───────────────────────────────────────────────────────────

def build_flight_fields(code: str, data: dict, aircraft_record_id: str | None) -> dict:
    dates = data["dates"]
    first = min(dates).strftime("%Y-%m-%d")
    last  = max(dates).strftime("%Y-%m-%d")

    fields: dict = {
        "Import Flight Key":  f"LEGACY|{code}",
        "Legacy Summary":     True,
        "Flight Date":        last,
        "Block Time":         round(data["total"], 1),
        "Night Time":         round(data["night_time"], 1),
        "Cross Country Time": round(data["xc"], 1),
        "Day Landing":        int(round(data["day_ldg"])),
        "Night Landing":      int(round(data["night_ldg"])),
        "Instrument Time":    round(data["act_in"], 1),
        "PIC Time":           round(data["pic"], 1),
        "SIC Time":           round(data["sic"], 1),
        "Dual Received":      round(data["dual"], 1),
        "Dual Given":         round(data["cfi"], 1),
        "Notes": (
            f"Legacy import from {SOURCE_FILENAME}\n"
            f"Date range: {first} – {last}\n"
            f"Source rows: {data['rows']}"
        ),
    }
    if aircraft_record_id:
        fields["Aircraft"] = [aircraft_record_id]
    return fields


def build_aircraft_fields(code: str) -> dict | None:
    meta = AIRCRAFT_META.get(code)
    if meta is None:
        return None
    faa_type, category, class_, engine = meta
    fields: dict = {
        "Aircraft":    code,
        "FAA Type":    faa_type,
        "Category":    category,
        "Engine Type": engine,
    }
    if class_ is not None:
        fields["Class"] = class_
    return fields


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import legacy logbook data as per-aircraft-type summary rows into Airtable.\n"
            "Dry-run by default — pass --commit to write."
        )
    )
    parser.add_argument(
        "--xlsx", type=Path, default=DEFAULT_XLSX, metavar="PATH",
        help=f"Path to AnytimeLogbooks .xlsx export (default: {DEFAULT_XLSX.name})",
    )
    parser.add_argument(
        "--commit", action="store_true",
        help="Write to Airtable. Omit to perform a dry-run.",
    )
    args = parser.parse_args()

    dry_run = not args.commit
    tag = "[DRY RUN] " if dry_run else ""

    if not args.xlsx.exists():
        sys.exit(f"File not found: {args.xlsx}")

    load_dotenv(ENV_PATH)
    api_key = os.environ.get("AIRTABLE_API_KEY", "").strip()
    if not api_key:
        sys.exit(f"AIRTABLE_API_KEY not set. Add it to {ENV_PATH}")

    print(f"{tag}Loading {args.xlsx.name} ...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb["Master"]
    aggregated = aggregate_master(ws)
    total_rows = sum(d["rows"] for d in aggregated.values())
    total_hrs  = sum(d["total"] for d in aggregated.values())
    print(f"  {len(aggregated)} aircraft types · {total_rows} source rows · {total_hrs:.1f} total hours\n")

    api = Api(api_key)
    flights_tbl  = api.table(BASE_ID, TBL_FLIGHTS)
    aircraft_tbl = api.table(BASE_ID, TBL_AIRCRAFT)

    existing_ac: dict[str, str] = {
        r["fields"].get("Aircraft", ""): r["id"]
        for r in aircraft_tbl.all(fields=["Aircraft"])
        if r["fields"].get("Aircraft")
    }
    print(f"Aircraft table: {len(existing_ac)} existing records.\n")

    created_ac = 0
    upserted_flights = 0

    for code in sorted(aggregated, key=lambda x: -aggregated[x]["total"]):
        data = aggregated[code]
        print(f"── {code} ({'%.1f' % data['total']} hrs) ──────────────────────")

        # ── Aircraft record ───────────────────────────────────────────────────
        aircraft_id = existing_ac.get(code)
        if aircraft_id:
            print(f"  Aircraft: existing ({aircraft_id})")
        else:
            ac_fields = build_aircraft_fields(code)
            if ac_fields is None:
                print(f"  [WARN] No metadata defined for {code} — skipping (add to AIRCRAFT_META).")
                print()
                continue
            meta = AIRCRAFT_META[code]
            print(
                f"  Aircraft: NEW — FAA Type={meta[0]}, Category={meta[1]}, "
                f"Class={meta[2]!r}, Engine={meta[3]}"
            )
            if not dry_run:
                rec = aircraft_tbl.create(ac_fields, typecast=True)
                aircraft_id = rec["id"]
                existing_ac[code] = aircraft_id
                created_ac += 1
                print(f"             → Created {aircraft_id}")

        # ── Flights upsert ────────────────────────────────────────────────────
        flight_fields = build_flight_fields(code, data, aircraft_id if not dry_run else None)
        dates = data["dates"]
        print(
            f"  Flights:  LEGACY|{code}  date={max(dates).strftime('%Y-%m-%d')}"
            f"  block={flight_fields['Block Time']}  PIC={flight_fields['PIC Time']}"
            f"  SIC={flight_fields['SIC Time']}  night={flight_fields['Night Time']}"
            f"  IMC={flight_fields['Instrument Time']}  XC={flight_fields['Cross Country Time']}"
            f"  dayLdg={flight_fields['Day Landing']}  ntLdg={flight_fields['Night Landing']}"
            f"  dualRcv={flight_fields['Dual Received']}  dualGvn={flight_fields['Dual Given']}"
        )
        if not dry_run:
            result = flights_tbl.upsert(
                [{"fields": flight_fields}],
                key_fields=["Import Flight Key"],
                typecast=True,
            )
            n_created = len(result.get("createdRecords", []))
            n_updated = len(result.get("updatedRecords", []))
            action = "Created" if n_created else "Updated"
            upserted_flights += 1
            print(f"             → {action} Flights record")
        print()

    # ── Summary ───────────────────────────────────────────────────────────────
    print("=" * 60)
    if dry_run:
        print("DRY RUN complete — no data was written.")
        print("Re-run with --commit to execute the import.")
    else:
        print(f"Import complete.")
        print(f"  Aircraft records created : {created_ac}")
        print(f"  Flights records upserted : {upserted_flights}")


if __name__ == "__main__":
    main()
